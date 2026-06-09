"""
fast_rate_capture_log.py
Headless capture test: run the radar at the OLD fast frame rate with NO display,
and log per-frame timing + loss to an Excel sheet.

PURPOSE
  The display lag was a display-side bottleneck (matplotlib), not a capture
  problem. This proves it: strip the display, keep the fast firmware rate, and
  show the capture path holds the rate cleanly with zero loss.

BEFORE RUNNING
  1. In firmware, set the FAST rate back on (common.c, the commented line):
         frameCfg.framePeriodicity = 50 * 1000000 / 5;   // 50 ms = 20 fps
     Rebuild + reload. (This script does NOT set the rate; the firmware does.)
  2. pip install openpyxl
  3. Start THIS script first, then resume the R5F core in CCS (same as always).

It runs for RUN_SECONDS (or until Ctrl+C), then writes the .xlsx and exits.
"""

import socket
import struct
import time

# ============================================================
# CONFIG - must match firmware / live_radar_tdma.py
# ============================================================
SAMPLES_PER_CHIRP = 256
NUM_RX            = 8
NUM_TX            = 6
NUM_LOOPS         = 64
CHIRPS_PER_FRAME  = NUM_LOOPS * NUM_TX          # 384

EXPECTED_DT_MS = 50.0      # the OLD fast rate we're testing (20 fps)
RUN_SECONDS    = 30        # how long to capture before writing the sheet
OUT_XLSX       = "fast_rate_capture.xlsx"

# ---- header format (must match firmware) ----
HEADER_MAGIC = 0xA1B2C3D4
HEADER_BYTES = 16
MAGIC_LE     = struct.pack("<I", HEADER_MAGIC)
BYTES_PER_SAMPLE = 4
DATA_PER_CHIRP   = SAMPLES_PER_CHIRP * NUM_RX * BYTES_PER_SAMPLE   # 8192
BLOCK_BYTES      = DATA_PER_CHIRP + HEADER_BYTES                   # 8208

# ---- network (same as the capture script) ----
FPGA_IP, HOST_IP = "192.168.33.180", "192.168.33.30"
CONFIG_PORT, DATA_PORT = 4096, 4098
DCA_HEADER_BYTES = 10
SOCKET_RECV_BUF = 2 ** 26
CMD_START_RECORD, CMD_STOP_RECORD = 0x05, 0x06


def dca_command(code, data=b""):
    return (struct.pack("<H", 0xA55A) + struct.pack("<H", code)
            + struct.pack("<H", len(data)) + data + struct.pack("<H", 0xEEAA))


def send_config_command(code):
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.bind((HOST_IP, CONFIG_PORT)); s.settimeout(2.0)
    s.sendto(dca_command(code), (FPGA_IP, CONFIG_PORT)); s.close()


# ============================================================
# Capture loop - light, no FFT, no display
# ============================================================
def capture():
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.setsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF, SOCKET_RECV_BUF)
    sock.bind((HOST_IP, DATA_PORT)); sock.settimeout(5.0)
    print(f"Listening on {HOST_IP}:{DATA_PORT}")

    send_config_command(CMD_START_RECORD)
    print("Streaming started. Resume the R5F core in CCS now.")
    print(f"Capturing for {RUN_SECONDS} s (Ctrl+C to stop early)...\n")

    rows = []                 # per-frame records for the sheet
    buf = bytearray()
    scan = 0

    cur_frame = None
    got = 0
    prev_gcount = None
    chirps_dropped = 0
    dca_pkts = dca_lost = 0
    expected_seq = None
    prev_bt = None
    t_end = None

    def finalize(frame_id, got_count):
        dt = (time.perf_counter() - prev_bt) * 1e3 if prev_bt is not None else 0.0
        rows.append({
            "frame_id": frame_id,
            "dt_ms": dt,
            "got": got_count,
            "missing": CHIRPS_PER_FRAME - got_count,
            "dropped_total": chirps_dropped,
            "dca_pkts": dca_pkts,
            "dca_lost": dca_lost,
        })

    try:
        while True:
            try:
                pkt = sock.recv(2048)
            except socket.timeout:
                print("No data 5s - waiting..."); continue
            if len(pkt) <= DCA_HEADER_BYTES:
                continue

            if t_end is None:                       # start clock on first real packet
                t_end = time.perf_counter() + RUN_SECONDS
            seq = struct.unpack("<I", pkt[0:4])[0]
            dca_pkts += 1
            if expected_seq is not None and seq > expected_seq:
                dca_lost += seq - expected_seq
            expected_seq = seq + 1

            buf.extend(pkt[DCA_HEADER_BYTES:])

            idx = buf.find(MAGIC_LE, scan)
            while idx != -1:
                if idx + HEADER_BYTES > len(buf):
                    break
                if idx >= DATA_PER_CHIRP:
                    _, gcount, fid, cid = struct.unpack("<4I", buf[idx:idx + HEADER_BYTES])

                    if prev_gcount is not None and gcount > prev_gcount + 1:
                        chirps_dropped += gcount - prev_gcount - 1
                    prev_gcount = gcount

                    if fid != cur_frame:
                        t = time.perf_counter()
                        if cur_frame is not None:
                            finalize(cur_frame, got)
                        prev_bt = t
                        cur_frame = fid
                        got = 0
                    got += 1

                scan = idx + HEADER_BYTES
                idx = buf.find(MAGIC_LE, scan)

            if scan > 0:
                del buf[:scan]
                scan = 0

            if t_end is not None and time.perf_counter() >= t_end:
                break
    except KeyboardInterrupt:
        print("Stopped early by user.")
    finally:
        send_config_command(CMD_STOP_RECORD)
        sock.close()

    print(f"Capture done: {len(rows)} frames recorded.")
    return rows


# ============================================================
# Excel report
# ============================================================
def write_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    bold = Font(name="Arial", bold=True)
    plain = Font(name="Arial")
    hdr_fill = PatternFill("solid", start_color="DDDDDD")
    center = Alignment(horizontal="center")

    wb = Workbook()

    # ---- Sheet 1: per-frame data ----
    ws = wb.active
    ws.title = "Frames"
    cols = ["frame_id", "dt_ms", "chirps_got", "missing",
            "chirps_dropped_total", "dca_pkts", "dca_lost"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold; cell.fill = hdr_fill; cell.alignment = center

    for r in rows:
        ws.append([r["frame_id"], round(r["dt_ms"], 2), r["got"], r["missing"],
                   r["dropped_total"], r["dca_pkts"], r["dca_lost"]])
    for col, w in zip("ABCDEFG", [12, 10, 12, 10, 20, 12, 12]):
        ws.column_dimensions[col].width = w

    last = len(rows) + 1                      # last data row (row 1 is header)

    # ---- Sheet 2: summary (formulas reference the Frames sheet) ----
    s = wb.create_sheet("Summary")
    s.column_dimensions["A"].width = 32
    s.column_dimensions["B"].width = 18

    def put(row, label, value, is_formula=False, bold_lbl=False):
        a = s.cell(row=row, column=1, value=label)
        a.font = bold if bold_lbl else plain
        b = s.cell(row=row, column=2, value=value)
        b.font = plain

    put(1, "Fast-rate capture summary", None, bold_lbl=True)
    put(3, "Target rate (firmware)", f"{EXPECTED_DT_MS:.0f} ms  ({1000/EXPECTED_DT_MS:.0f} fps)")
    put(4, "Frames captured", f"=COUNT(Frames!A2:A{last})")
    # dt stats skip row 2 (first frame's dt is 0, no prior boundary)
    put(5, "Mean dt (ms)", f"=AVERAGE(Frames!B3:B{last})")
    put(6, "Min dt (ms)",  f"=MIN(Frames!B3:B{last})")
    put(7, "Max dt (ms)",  f"=MAX(Frames!B3:B{last})")
    put(8, "Effective fps", "=1000/B5")
    put(9, "Total chirps dropped", f"=MAX(Frames!E2:E{last})")
    put(10, "Total DCA packets", f"=MAX(Frames!F2:F{last})")
    put(11, "Total DCA lost", f"=MAX(Frames!G2:G{last})")
    put(12, "DCA loss %", "=IF(B10=0,0,100*B11/B10)")
    put(13, "Frames with missing chirps", f"=COUNTIF(Frames!D2:D{last},\">0\")")
    # verdict: rate within 20% of target AND zero loss/drops/missing
    put(15, "VERDICT", (
        f'=IF(AND(ABS(B5-{EXPECTED_DT_MS})<{0.2*EXPECTED_DT_MS},'
        f'B9=0,B11=0,B13=0),'
        f'"PASS - holds {EXPECTED_DT_MS:.0f}ms with zero loss",'
        f'"CHECK - see stats above")'), bold_lbl=True)
    s.cell(row=15, column=2).font = bold

    wb.save(path)
    print(f"Wrote {path}")


def main():
    rows = capture()
    if not rows:
        print("No frames captured - nothing to write.")
        return
    write_xlsx(rows, OUT_XLSX)


if __name__ == "__main__":
    main()
